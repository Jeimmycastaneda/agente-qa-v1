import io
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader

from agente_qa.export.excel import create_excel
from agente_qa.export.pdf import create_pdf
from agente_qa.integrations.azure_runtime import _azure_steps_xml
from config.qa_config import AZURE_COLUMNS, MATRIZ_COLUMNS


def sample_data():
    return {
        "PRODUCT": "Cotizadores Web",
        "TEST_CASES": [
            {
                "ID": "CP-AC-CAC-00001",
                "Title": "Consultar cotización en estado Cotizado",
                "Product": "Cotizadores Web",
                "Module": "Cotizador Autos Colectivos",
                "Description": "Validar consulta de cotización",
                "Expected Result": "Se muestra la información de la cotización",
                "Preconditions": "Usuario autenticado",
                "Related Use Case": "CU-325",
                "Scenario": "Consulta de cotización",
                "Steps": [
                    {"Step #": 1, "Action": "Ingresar al módulo", "Expected": "Se muestra la pantalla principal"},
                    {"Step #": 2, "Action": "Consultar cotización", "Expected value": "Se muestran los datos"},
                ],
            }
        ],
        "USE_CASES": [{"ID": "CU-325", "Name": "Consultar cotización"}],
    }


def test_excel_preserves_approved_columns_and_required_defaults():
    content = create_excel(sample_data(), "Autos Colectivos")
    workbook = load_workbook(io.BytesIO(content), data_only=True)

    assert workbook.sheetnames == ["Azure Import", "Matriz QA"]
    assert [cell.value for cell in workbook["Azure Import"][1]] == AZURE_COLUMNS
    assert [cell.value for cell in workbook["Matriz QA"][1]] == MATRIZ_COLUMNS

    rows = list(workbook["Azure Import"].iter_rows(min_row=2, values_only=True))
    assert rows[0][7] == "COTIZADORES WEB\\DESARROLLO"
    assert rows[0][9] == "Proyecto"
    assert rows[1][6] == "Se muestra la pantalla principal"
    assert rows[2][6] == "Se muestran los datos"


def test_excel_uses_selected_suite_name_for_case_id():
    data = sample_data()
    data["SUITE_NAME"] = "DRC"

    content = create_excel(data, "Autos Colectivos")
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    rows = list(workbook["Azure Import"].iter_rows(min_row=2, values_only=True))

    assert rows[0][3].startswith("CP-AUDRC-00001 ")
    assert rows[0][3] == "CP-AUDRC-00001 Consultar cotización en estado Cotizado"
    assert rows[0][3].count("|") == 0


def test_excel_uses_sequential_ids_when_model_returns_duplicate_ids():
    data = sample_data()
    data["SUITE_NAME"] = "DRC"
    data["TEST_CASES"] = [
        dict(data["TEST_CASES"][0]),
        {**data["TEST_CASES"][0], "Title": "Validar segunda consulta", "Related Use Case": "CU-326"},
    ]
    data["TEST_CASES"][0]["ID"] = "CP-AU-99999"
    data["TEST_CASES"][1]["ID"] = "CP-AU-99999"

    content = create_excel(data, "Autos Colectivos")
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    rows = list(workbook["Azure Import"].iter_rows(min_row=2, values_only=True))

    assert rows[0][3].startswith("CP-AUDRC-00001 ")
    assert rows[3][3].startswith("CP-AUDRC-00002 ")


def test_excel_uses_ho_when_suite_data_identifies_hogar():
    data = sample_data()
    data["SUITE_NAME"] = "DRC"
    data["TEST_CASES"][0]["Module"] = "Hogar"
    data["TEST_CASES"][0]["Title"] = "Validar cobertura de vivienda"

    content = create_excel(data, "Autos Colectivos")
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    rows = list(workbook["Azure Import"].iter_rows(min_row=2, values_only=True))

    assert rows[0][3] == "CP-HODRC-00001 Validar cobertura de vivienda"


def test_pdf_export_generates_readable_pdf_bytes():
    content = create_pdf(sample_data(), "Autos Colectivos", "HU-TEST")
    assert content.startswith(b"%PDF-")
    assert len(content) > 1000


def test_pdf_preserves_expected_when_step_uses_either_key():
    data = sample_data()
    data["TEST_CASES"][0]["Steps"] = [
        {"Step #": 1, "Action": "Paso con Expected", "Expected": "Resultado A"},
        {"Step #": 2, "Action": "Paso con Expected value", "Expected value": "Resultado B"},
    ]
    content = create_pdf(data, "Autos Colectivos", "HU-TEST")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages)
    assert "Resultado A" in text
    assert "Resultado B" in text


def test_runtime_steps_xml_uses_action_and_expected_fields():
    result = _azure_steps_xml(sample_data()["TEST_CASES"][0]["Steps"])
    root = ET.fromstring(result)
    steps = root.findall("step")

    assert len(steps) == 2
    values = steps[0].findall("parameterizedString")
    assert values[0].text == "Ingresar al módulo"
    assert values[1].text == "Se muestra la pantalla principal"
